import openpyxl
from PyQt5.QtCore import QObject
from global_defs import *
import log_utils
import os
import string


import datetime

log = log_utils.logging_init(__file__)

machine_serial_number_column=1
cpu_serial_number_column=2
eth_mac_addr_column=3
wlan_mac_addr_column=4
ledclient_sw_version_column=5
ledserver_sw_version_column=6
ledsystem_sw_version_column=7



class XlsMod(QObject):
	def __init__(self, file_uri, parent):
		super().__init__(parent)
		self.file_uri = file_uri
		log.debug("%s:", datetime.datetime.now().strftime("%Y-%m-%d"))
		self.str_date_now = datetime.datetime.now().strftime("%Y-%m-%d")
		if os.path.exists(self.file_uri) is False:
			wb = openpyxl.Workbook()
			# wb.create_sheet(self.str_date_now)
			wb.save(self.file_uri)
		self.xls_fd = self.load_xls()
		self.work_sheet = self.get_today_sheet() # self.xls_fd._sheets[1]
		log.debug("self.xls_fd._sheets title = %s", self.work_sheet.title)
		log.debug("self.xls_fd._sheets len = %d", len(self.xls_fd._sheets))
		log.debug("self.work_sheet.max_row = %d", self.work_sheet.max_row)
		log.debug("self.work_sheet.max_column = %d", self.work_sheet.max_column)
		self.working_row = self.work_sheet.max_row + 1


	def load_xls(self):
		# 讀取 Excel 檔案
		wb = openpyxl.load_workbook(self.file_uri)
		return wb

	def get_sheets(self):
		return self.xls_fd.get_sheet_names()

	def get_active_sheet(self):
		return self.xls_fd.active()

	def get_today_sheet(self):
		log.debug("A self.xls_fd._sheets len = %d", len(self.xls_fd._sheets))
		for i in range(len(self.xls_fd._sheets)):
			if self.xls_fd._sheets[i].title == self.str_date_now:
				return self.xls_fd._sheets[i]
		self.xls_fd.create_sheet(self.str_date_now)
		self.init_work_sheet(self.xls_fd._sheets[len(self.xls_fd._sheets) - 1])
		log.debug("B self.xls_fd._sheets len = %d", len(self.xls_fd._sheets))
		self.xls_fd.save(self.file_uri)
		return self.xls_fd._sheets[len(self.xls_fd._sheets) - 1]

	def save_work_sheet(self):
		log.debug("save_work_sheet")
		self.xls_fd.save(self.file_uri)

	def init_work_sheet(self, ws):
		ws.cell(column=1, row=1).value = "SerialNumber"
		ws.cell(column=2, row=1).value = "CPU_SerialNumber"
		ws.cell(column=3, row=1).value = "ETH MAC ADDR"
		ws.cell(column=4, row=1).value = "WLAN MAC ADDR"
		ws.cell(column=5, row=1).value = "LedClient SW Version"
		ws.cell(column=6, row=1).value = "LedServer SW Version"
		ws.cell(column=7, row=1).value = "LedSystem SW Version"

		for i in range(ord('A'), ord('H') + 1):
			ws.column_dimensions[str(chr(i))].width = 20
		log.debug("ws.max_row : %d", ws.max_row)

	def write_cpu_serial_number(self, cpu_serial_number):
		self.work_sheet.cell(column=cpu_serial_number_column, row=self.working_row).value = cpu_serial_number

	def write_box_serial_number(self, box_serial_number):
		self.work_sheet.cell(column=machine_serial_number_column, row=self.working_row).value = box_serial_number

	def write_eth_mac_address(self, eth_mac_address):
		self.work_sheet.cell(column=eth_mac_addr_column, row=self.working_row).value = eth_mac_address

	def write_wlan_mac_address(self, wlan_mac_address):
		self.work_sheet.cell(column=wlan_mac_addr_column, row=self.working_row).value = wlan_mac_address

	def write_ledclient_sw_version(self, ledclient_sw_version):
		self.work_sheet.cell(column=ledclient_sw_version_column, row=self.working_row).value = ledclient_sw_version

	def write_ledserver_sw_version(self, ledserver_sw_version):
		self.work_sheet.cell(column=ledserver_sw_version_column, row=self.working_row).value = ledserver_sw_version

	def write_ledsystem_sw_version(self, ledsystem_sw_version):
		self.work_sheet.cell(column=ledsystem_sw_version_column, row=self.working_row).value = ledsystem_sw_version

	def increase_working_row(self):
		self.working_row += 1


